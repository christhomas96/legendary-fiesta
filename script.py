import openpyxl
import smtplib
import time
from twilio.rest import Client
from calendly import CalendlyAPI
from datetime import datetime, timedelta

# Replace with your Twilio account SID and auth token
TWILIO_ACCOUNT_SID = 'YOUR_TWILIO_ACCOUNT_SID'
TWILIO_AUTH_TOKEN = 'YOUR_TWILIO_AUTH_TOKEN'
TWILIO_PHONE_NUMBER = 'YOUR_TWILIO_PHONE_NUMBER'
TWILIO_WHATSAPP_NUMBER = 'whatsapp:YOUR_TWILIO_WHATSAPP_NUMBER'

# Replace with your Calendly API key
CALENDLY_API_KEY = 'YOUR_CALENDLY_API_KEY'

# Initialize the Calendly API client
calendly = CalendlyAPI(CALENDLY_API_KEY)

# Load the Excel file
workbook = openpyxl.load_workbook('data.xlsx')
sheet = workbook.active

# Iterate through each row in the Excel sheet
for row in range(2, sheet.max_row + 1):
    company_name = sheet.cell(row=row, column=1).value
    calendly_link = sheet.cell(row=row, column=2).value
    interviewee_email = sheet.cell(row=row, column=3).value
    interviewee_phone = sheet.cell(row=row, column=4).value
    company_contact = sheet.cell(row=row, column=5).value  # Assuming company contact is in column 5

    # Extract the Calendly event ID and expiration date from the link
    event_id, expiration_date = extract_calendly_info(calendly_link)

    # Check if the Calendly link is about to expire
    if expiration_date - datetime.now() < timedelta(days=7):
        notify_company_link_expiration(company_contact, company_name, calendly_link)

    # Send the Calendly link to the interviewee
    send_calendly_link(interviewee_email, interviewee_phone, company_name, calendly_link)

    # Check if the interviewee scheduled a meeting
    if check_scheduled_meeting(interviewee_email, interviewee_phone, company_name, calendly_link):
        print(f"Interviewee {interviewee_email}/{interviewee_phone} scheduled a meeting for {company_name}")
    else:
        # Send follow-up communications
        follow_up(interviewee_email, interviewee_phone, company_name, calendly_link)

        # Manual intervention if needed
        manual_intervention(interviewee_email, interviewee_phone, company_name)

    # Monitor Calendly event updates for the company
    monitor_calendly_events(event_id, company_contact, company_name)

    # Follow up with the company for unscheduled interviews
    follow_up_company_unscheduled(company_contact, company_name, event_id)

# Function to extract Calendly event ID and expiration date from the link
def extract_calendly_info(calendly_link):
    # Implement logic to extract the event ID and expiration date from the Calendly link
    # This logic may vary depending on the link structure provided by Calendly
    event_id = calendly_link.split('/')[-1]
    expiration_date = datetime(2023, 8, 1)  # Replace with actual expiration date extraction logic
    return event_id, expiration_date

# Function to notify the company about an expiring Calendly link
def notify_company_link_expiration(company_contact, company_name, calendly_link):
    message = f"Dear {company_name},\n\nYour Calendly link ({calendly_link}) is about to expire. Please update or extend the link's validity to avoid any disruptions in the scheduling process.\n\nBest regards,\nYour Team"
    send_email(company_contact, message)

# Function to monitor Calendly event updates for the company
def monitor_calendly_events(event_id, company_contact, company_name):
    # Query the Calendly API for event updates (cancellations, rescheduling, etc.)
    event_updates = calendly.event_updates.list(event_id=event_id)

    # Process event updates and notify the company and/or interviewees as needed
    for update in event_updates:
        if update.event_status == 'canceled':
            notify_interviewee_event_canceled(update.invitee_email, company_name)
            # ... (additional logic for handling event cancellations)

        elif update.event_status == 'rescheduled':
            notify_interviewee_event_rescheduled(update.invitee_email, company_name, update.new_event_time)
            # ... (additional logic for handling event rescheduling)

        elif update.event_status == 'scheduled':
            notify_company_interviewer(company_contact, company_name, update.event_time, update.invitee_email)

# Function to follow up with the company for unscheduled interviews
def follow_up_company_unscheduled(company_contact, company_name, event_id):
    # Check for unscheduled interviews after a certain period (e.g., one week)
    if datetime.now() - last_follow_up_time > timedelta(days=7):
        unscheduled_interviews = calendly.scheduled_events.list(event_id=event_id, status='unscheduled')
        if unscheduled_interviews:
            message = f"Dear {company_name},\n\nWe noticed that {len(unscheduled_interviews)} interviews remain unscheduled for your company. Please review and update your Calendly link or provide additional information to facilitate the scheduling process.\n\nBest regards,\nYour Team"
            send_email(company_contact, message)
            last_follow_up_time = datetime.now()

# Function to send the Calendly link to the interviewee
def send_calendly_link(email, phone, company_name, calendly_link):
    message = f"Hello, please use the following Calendly link to schedule an interview with {company_name}:\n{calendly_link}"

    try:
        send_whatsapp(phone, message)
    except Exception as e:
        print(f"Error sending WhatsApp message: {e}")
        try:
            make_call(phone, message)
        except Exception as e:
            print(f"Error making call: {e}")
            send_email(email, message)

# Function to check if the interviewee scheduled a meeting
def check_scheduled_meeting(email, phone, company_name, calendly_link):
    # Extract the Calendly event ID from the link
    event_id = calendly_link.split('/')[-1]

    # Query the Calendly API for scheduled events
    scheduled_events = calendly.scheduled_events.list(event_id=event_id)

    # Check if any scheduled events match the interviewee's email or phone number
    for event in scheduled_events:
        if event.invitee_email == email or event.invitee_phone_number == phone:
            return True

    return False

# Function to send follow-up communications
def follow_up(email, phone, company_name, calendly_link):
    message = f"This is a friendly reminder to schedule an interview with {company_name} using the following Calendly link:\n{calendly_link}"

    try:
        send_whatsapp(phone, message)
        send_email(email, message)
    except Exception as e:
        print(f"Error sending follow-up communication: {e}")

    time.sleep(10800)  # Wait for 3 hours

    if not check_scheduled_meeting(email, phone, company_name, calendly_link):
        print(f"Interviewee {email}/{phone} did not schedule a meeting for {company_name}. Escalating to manual intervention.")
        manual_intervention(email, phone, company_name)

# Function for manual intervention
def manual_intervention(email, phone, company_name):
    print(f"Manual intervention required for interviewee {email}/{phone} for {company_name}")
    # Your code for manual intervention (e.g., assigning the case to a team member)
    team_member = assign_case_to_team_member(email, phone, company_name)
    if team_member:
        print(f"Case assigned to {team_member} for manual follow-up.")
    else:
        print("Unable to assign case for manual follow-up.")

# Function to send an email
def send_email(recipient, message):
    # Replace with your email credentials
    sender_email = "YOUR_EMAIL_ADDRESS"
    sender_password = "YOUR_EMAIL_PASSWORD"

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient, message)
        print(f"Email sent to {recipient}")
    except Exception as e:
        print(f"Error sending email: {e}")
    finally:
        server.quit()

# Function to make a call
def make_call(phone_number, message):
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    try:
        call = client.calls.create(
            twiml=f'<Response><Say>{message}</Say></Response>',
            to=phone_number,
            from_=TWILIO_PHONE_NUMBER
        )
        print(f"Call made to {phone_number}")
    except Exception as e:
        print(f"Error making call: {e}")

# Function to send a WhatsApp message
def send_whatsapp(phone_number, message):
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    try:
        message = client.messages.create(
            body=message,
            from_=TWILIO_WHATSAPP_NUMBER,
            to=f"whatsapp:{phone_number}"
        )
        print(f"WhatsApp message sent to {phone_number}")
    except Exception as e:
        print(f"Error sending WhatsApp message: {e}")

# Function to notify interviewee about event cancellation
def notify_interviewee_event_canceled(invitee_email, company_name):
    message = f"Dear Interviewee,\n\nWe regret to inform you that the interview with {company_name} has been canceled. Please stay tuned for further updates.\n\nBest regards,\nYour Team"
    send_email(invitee_email, message)

# Function to notify interviewee about event rescheduling
def notify_interviewee_event_rescheduled(invitee_email, company_name, new_event_time):
    message = f"Dear Interviewee,\n\nYour interview with {company_name} has been rescheduled to {new_event_time}. Please make a note of the new date and time.\n\nBest regards,\nYour Team"
    send_email(invitee_email, message)

# Function to notify company interviewer about scheduled interview
def notify_company_interviewer(company_contact, company_name, event_time, invitee_email):
    message = f"Dear Company Interviewer,\n\nAn interview has been scheduled with {company_name} for {invitee_email} at {event_time}. Please make a note of this and follow up with the interviewee to ensure their attendance.\n\nBest regards,\nYour Team"
    send_email(company_contact, message)

# Function to assign case to a team member for manual follow-up (placeholder)
def assign_case_to_team_member(email, phone, company_name):
    # Implement your logic to assign the case to a team member
    # For example, you could use a round-robin or load balancing strategy
    # Return the team member's name or identifier
    return "Team Member 1"